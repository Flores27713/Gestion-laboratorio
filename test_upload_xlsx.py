"""
Test de upload EXCEL (.xlsx) contra el servidor LIVE.
"""
import urllib.request
import json
import io
import openpyxl

BASE = "https://gestion-laboratorio-98ze.onrender.com"

def api_get(path):
    req = urllib.request.Request(f"{BASE}{path}")
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))

def api_upload(filename, file_bytes, content_type):
    boundary = "----WebKitFormBoundary7MA4YWxkTrZu0gW"
    body = (
        f"--{boundary}\r\n"
        f'Content-Disposition: form-data; name="file"; filename="{filename}"\r\n'
        f"Content-Type: {content_type}\r\n\r\n"
    ).encode("utf-8") + file_bytes + f"\r\n--{boundary}--\r\n".encode("utf-8")
    
    req = urllib.request.Request(
        f"{BASE}/api/inventory/upload-csv",
        data=body,
        headers={"Content-Type": f"multipart/form-data; boundary={boundary}"},
        method="POST"
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read().decode("utf-8"))

# STEP 1: Current state
print("=== PASO 1: Estado actual ===")
items = api_get("/api/inventory?search=DISPLAY")
item1 = [i for i in items if i["id"] == 1][0]
print(f"  DISPLAY: stock={item1['stock']}, avail={item1['avail']}")

# STEP 2: Create and upload XLSX
print("\n=== PASO 2: Creando y subiendo archivo EXCEL (.xlsx) con stock=777 ===")
wb = openpyxl.Workbook()
ws = wb.active
ws.append(["ID", "Nombre del Artículo", "Ubicación Bodega", "Stock Total", "En Uso / Prestado", "En Mal Estado", "Stock Disponible", "Centro de Costo"])
ws.append([1, "DISPLAY", "BO-EST-01", 777, 0, 0, 777, "Informática"])

xlsx_buf = io.BytesIO()
wb.save(xlsx_buf)
xlsx_bytes = xlsx_buf.getvalue()
print(f"  XLSX size: {len(xlsx_bytes)} bytes")

result = api_upload("test_inventario.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
print(f"  Response: {result}")

# STEP 3: Verify
print("\n=== PASO 3: Verificando ===")
items = api_get("/api/inventory?search=DISPLAY")
item1_after = [i for i in items if i["id"] == 1][0]
print(f"  DISPLAY ahora: stock={item1_after['stock']}, avail={item1_after['avail']}")

if item1_after["stock"] == 777:
    print("  [OK] UPLOAD EXCEL FUNCIONA")
else:
    print(f"  [FAIL] NO FUNCIONA - stock sigue en {item1_after['stock']}")

# STEP 4: Restore
print("\n=== PASO 4: Restaurando ===")
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.append(["ID", "Nombre del Artículo", "Ubicación Bodega", "Stock Total", "En Uso / Prestado", "En Mal Estado", "Stock Disponible", "Centro de Costo"])
ws2.append([1, "DISPLAY", "BO-EST-01", 23, 0, 0, 23, "Informática"])
xlsx_buf2 = io.BytesIO()
wb2.save(xlsx_buf2)
result2 = api_upload("restore.xlsx", xlsx_buf2.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
print(f"  Restore: {result2}")

# Now test with the SAME structure as the user's real file "Inventario_2026_INOVALAB.xlsx"
# The user's file might have different column order or names
print("\n=== PASO 5: Probando con estructura alternativa (sin columna ID) ===")
wb3 = openpyxl.Workbook()
ws3 = wb3.active
# Maybe user's file has: Nombre, Ubicacion, Stock, En Uso, Mal Estado, Disponible, Centro Costo (no ID column)
ws3.append(["Nombre Artículo", "Ubicación Bodega", "Stock Total", "En Uso", "Mal Estado", "Disponible", "Centro de Costo"])
ws3.append(["DISPLAY", "BO-EST-01", 555, 0, 0, 555, "Informática"])
xlsx_buf3 = io.BytesIO()
wb3.save(xlsx_buf3)
result3 = api_upload("alt_format.xlsx", xlsx_buf3.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
print(f"  Response: {result3}")

items = api_get("/api/inventory?search=DISPLAY")
item1_alt = [i for i in items if i["id"] == 1][0]
print(f"  DISPLAY ahora: stock={item1_alt['stock']}, avail={item1_alt['avail']}")
if item1_alt["stock"] == 555:
    print("  [OK] Formato sin ID funciona")
else:
    print(f"  [FAIL] Formato sin ID no funciono - stock={item1_alt['stock']}")

# Final restore
wb4 = openpyxl.Workbook()
ws4 = wb4.active
ws4.append(["ID", "Nombre del Artículo", "Ubicación", "Stock Total", "En Uso", "Mal Estado", "Disponible", "Centro de Costo"])
ws4.append([1, "DISPLAY", "BO-EST-01", 23, 0, 0, 23, "Informática"])
xlsx_buf4 = io.BytesIO()
wb4.save(xlsx_buf4)
api_upload("final_restore.xlsx", xlsx_buf4.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

items = api_get("/api/inventory?search=DISPLAY")
item1_final = [i for i in items if i["id"] == 1][0]
print(f"\n  Final restore: DISPLAY stock={item1_final['stock']}, avail={item1_final['avail']}")
print("\n=== TEST COMPLETADO ===")
